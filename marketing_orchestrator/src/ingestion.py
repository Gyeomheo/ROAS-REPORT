"""Excel ingestion/output helpers with Polars-first and openpyxl fallback."""

from __future__ import annotations

import json
import os
from datetime import date
from pathlib import Path
from typing import Any, Dict, Sequence

import polars as pl


DIMENSIONS: list[str] = ["SUBSIDIARY", "CHANNEL", "DIVISION", "PRODUCT"]
RAW_METRIC_MAP: dict[str, str] = {
    "PLATFORM_SPEND_USD": "Spend",
    "GROSS_REVENUE": "Revenue",
    "PLATFORM_CLICKS": "Clicks",
    "GROSS_ORDERS": "Orders",
}
METRICS: list[str] = ["Spend", "Revenue", "Clicks", "Orders"]
OBJECTIVE_CANDIDATES: tuple[str, ...] = ("OBJECTIVE", "Objective", "objective")
TARGET_OBJECTIVE_VALUE = "CONVERSION"
TARGET_DIVISIONS: tuple[str, ...] = ("MX", "VD", "DA")
ENGINE_METRIC_COLUMNS: list[str] = [
    "Spend_curr",
    "Spend_prev",
    "Revenue_curr",
    "Revenue_prev",
    "Clicks_curr",
    "Clicks_prev",
    "Orders_curr",
    "Orders_prev",
]
ENGINE_COLUMNS: list[str] = [*DIMENSIONS, *ENGINE_METRIC_COLUMNS]
LONG_COLUMNS_YEAR_WITH_OBJECTIVE: list[str] = [*DIMENSIONS, *RAW_METRIC_MAP.keys(), "OBJECTIVE", "Year", "Month", "Day"]
LONG_COLUMNS_YEAR_ALT_WITH_OBJECTIVE: list[str] = [*DIMENSIONS, *RAW_METRIC_MAP.keys(), "OBJECTIVE", "YEAR", "Month", "Day"]
LONG_COLUMNS_YEAR_UPPER_WITH_OBJECTIVE: list[str] = [*DIMENSIONS, *RAW_METRIC_MAP.keys(), "OBJECTIVE", "YEAR", "MONTH", "DAY"]
LONG_COLUMNS_YEAR_MIN_WITH_OBJECTIVE: list[str] = [*DIMENSIONS, *RAW_METRIC_MAP.keys(), "OBJECTIVE", "Year"]
LONG_COLUMNS_YEAR_ALT_MIN_WITH_OBJECTIVE: list[str] = [*DIMENSIONS, *RAW_METRIC_MAP.keys(), "OBJECTIVE", "YEAR"]
LONG_COLUMNS_YEAR: list[str] = [*DIMENSIONS, *RAW_METRIC_MAP.keys(), "Year", "Month", "Day"]
LONG_COLUMNS_YEAR_ALT: list[str] = [*DIMENSIONS, *RAW_METRIC_MAP.keys(), "YEAR", "Month", "Day"]
LONG_COLUMNS_YEAR_UPPER: list[str] = [*DIMENSIONS, *RAW_METRIC_MAP.keys(), "YEAR", "MONTH", "DAY"]
LONG_COLUMNS_YEAR_MIN: list[str] = [*DIMENSIONS, *RAW_METRIC_MAP.keys(), "Year"]
LONG_COLUMNS_YEAR_ALT_MIN: list[str] = [*DIMENSIONS, *RAW_METRIC_MAP.keys(), "YEAR"]
PREFERRED_COLUMN_SETS: list[list[str]] = [
    LONG_COLUMNS_YEAR_WITH_OBJECTIVE,
    LONG_COLUMNS_YEAR_ALT_WITH_OBJECTIVE,
    LONG_COLUMNS_YEAR_UPPER_WITH_OBJECTIVE,
    LONG_COLUMNS_YEAR_MIN_WITH_OBJECTIVE,
    LONG_COLUMNS_YEAR_ALT_MIN_WITH_OBJECTIVE,
    LONG_COLUMNS_YEAR,
    LONG_COLUMNS_YEAR_ALT,
    LONG_COLUMNS_YEAR_UPPER,
    LONG_COLUMNS_YEAR_MIN,
    LONG_COLUMNS_YEAR_ALT_MIN,
    ENGINE_COLUMNS,
]
DEFAULT_CURR_YEAR = date.today().year
DEFAULT_PREV_YEAR = DEFAULT_CURR_YEAR - 1


def _parse_error_threshold() -> float:
    raw = os.getenv("ROAS_PARSE_ERROR_THRESHOLD", "0.01")
    try:
        threshold = float(raw)
    except ValueError as exc:
        raise ValueError(f"Invalid ROAS_PARSE_ERROR_THRESHOLD: {raw}") from exc
    if threshold < 0 or threshold > 1:
        raise ValueError(f"ROAS_PARSE_ERROR_THRESHOLD must be in [0, 1], got {threshold}")
    return threshold


METRIC_PARSE_ERROR_THRESHOLD = _parse_error_threshold()


def _import_openpyxl() -> tuple[Any, Any]:
    try:
        from openpyxl import Workbook, load_workbook
    except Exception as exc:  # pragma: no cover
        raise RuntimeError("openpyxl is required for Excel fallback I/O.") from exc
    return Workbook, load_workbook


def _normalize_headers(raw_headers: Sequence[Any]) -> list[str]:
    headers: list[str] = []
    seen: dict[str, int] = {}
    for idx, value in enumerate(raw_headers):
        base = str(value).strip() if value not in (None, "") else f"column_{idx + 1}"
        count = seen.get(base, 0)
        name = base if count == 0 else f"{base}_{count + 1}"
        seen[base] = count + 1
        headers.append(name)
    return headers


def _select_sheet_name(path: Path, preferred_sheet: str) -> str:
    try:
        _, load_workbook = _import_openpyxl()
    except Exception:
        return preferred_sheet

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet_names = list(workbook.sheetnames)
    workbook.close()
    if not sheet_names:
        raise ValueError(f"No sheets found in {path}")
    if preferred_sheet in sheet_names:
        return preferred_sheet
    return sheet_names[0]


def _read_excel_polars(path: Path, **kwargs: Any) -> Any:
    """Use larger schema sampling when supported to avoid dtype inference warnings."""
    try:
        return pl.read_excel(path, infer_schema_length=10000, **kwargs)  # type: ignore[arg-type]
    except TypeError:
        return pl.read_excel(path, **kwargs)  # type: ignore[arg-type]


def _frame_from_polars_result(frame: Any, preferred_sheet: str) -> pl.DataFrame:
    if isinstance(frame, dict):
        if preferred_sheet in frame:
            return frame[preferred_sheet]
        first_key = next(iter(frame.keys()), None)
        if first_key is None:
            return pl.DataFrame()
        return frame[first_key]
    return frame


def _read_with_polars(path: Path, preferred_sheet: str, target_sheet: str) -> pl.DataFrame:
    if not hasattr(pl, "read_excel"):
        raise RuntimeError("polars.read_excel is not available in this environment.")

    tried: list[str] = []
    for sheet_name in [target_sheet, preferred_sheet]:
        if not sheet_name or sheet_name in tried:
            continue
        tried.append(sheet_name)

        try:
            frame = _read_excel_polars(path, sheet_name=sheet_name)
            return _frame_from_polars_result(frame, preferred_sheet)
        except Exception:
            pass

        for columns in PREFERRED_COLUMN_SETS:
            try:
                frame = _read_excel_polars(path, sheet_name=sheet_name, columns=columns)
                return _frame_from_polars_result(frame, preferred_sheet)
            except Exception:
                continue

    try:
        frame = _read_excel_polars(path, sheet_name=preferred_sheet)
        return _frame_from_polars_result(frame, preferred_sheet)
    except Exception:
        frame = _read_excel_polars(path)
        return _frame_from_polars_result(frame, preferred_sheet)


def _read_with_openpyxl(path: Path, target_sheet: str, preferred_sheet: str) -> pl.DataFrame:
    _, load_workbook = _import_openpyxl()
    workbook = load_workbook(path, read_only=True, data_only=True)

    sheet_name = target_sheet
    if sheet_name not in workbook.sheetnames:
        sheet_name = preferred_sheet if preferred_sheet in workbook.sheetnames else workbook.sheetnames[0]

    worksheet = workbook[sheet_name]
    row_iter = worksheet.iter_rows(values_only=True)
    header_row = next(row_iter, None)
    if header_row is None:
        workbook.close()
        return pl.DataFrame()

    headers = _normalize_headers(header_row)
    records: list[dict[str, Any]] = []
    for values in row_iter:
        if values is None or all(value is None for value in values):
            continue
        row_data: dict[str, Any] = {}
        for idx, name in enumerate(headers):
            row_data[name] = values[idx] if idx < len(values) else None
        records.append(row_data)

    workbook.close()
    if not records:
        return pl.DataFrame({name: [] for name in headers})
    return pl.DataFrame(records)


def _year_expr(column_name: str) -> pl.Expr:
    return pl.coalesce(
        [
            pl.col(column_name).cast(pl.Int64, strict=False),
            pl.col(column_name)
            .cast(pl.Utf8, strict=False)
            .str.extract(r"(\d{4})", group_index=1)
            .cast(pl.Int64, strict=False),
        ]
    )


def _metric_text_expr(column_name: str) -> pl.Expr:
    return pl.col(column_name).cast(pl.Utf8, strict=False).str.strip_chars()


def _metric_parsed_expr(column_name: str) -> pl.Expr:
    return _metric_text_expr(column_name).str.replace_all(",", "").cast(pl.Float64, strict=False)


def _metric_parse_error_expr(column_name: str) -> pl.Expr:
    text_expr = _metric_text_expr(column_name)
    parsed_expr = _metric_parsed_expr(column_name)
    return (
        (text_expr.is_not_null() & (text_expr != "") & parsed_expr.is_null())
        .cast(pl.UInt32)
        .alias(f"__parse_error_{column_name}")
    )


def _metric_expr(column_name: str) -> pl.Expr:
    return _metric_parsed_expr(column_name).fill_null(0.0).alias(column_name)


def _validate_metric_parse_errors(
    df: pl.DataFrame,
    metric_columns: Sequence[str],
    context: str,
    threshold: float = METRIC_PARSE_ERROR_THRESHOLD,
) -> None:
    if df.is_empty() or threshold <= 0:
        return
    targets = [column for column in metric_columns if column in df.columns]
    if not targets:
        return

    checks_df = df.select([_metric_parse_error_expr(column) for column in targets])
    row_count = int(df.height)
    failures: list[str] = []
    for column in targets:
        check_col = f"__parse_error_{column}"
        count_value = checks_df.select(pl.col(check_col).sum()).to_series(0)[0]
        parse_error_count = int(count_value or 0)
        parse_error_ratio = parse_error_count / row_count if row_count > 0 else 0.0
        if parse_error_ratio > threshold:
            failures.append(f"{column}={parse_error_ratio:.2%} ({parse_error_count}/{row_count})")

    if failures:
        joined = ", ".join(failures)
        raise ValueError(
            f"Data quality check failed in {context}: metric parse error ratio exceeds {threshold:.2%} ({joined})"
        )


def _int_expr(column_name: str) -> pl.Expr:
    return pl.coalesce(
        [
            pl.col(column_name).cast(pl.Int64, strict=False),
            pl.col(column_name)
            .cast(pl.Utf8, strict=False)
            .str.extract(r"(\d+)", group_index=1)
            .cast(pl.Int64, strict=False),
        ]
    )


def _dimension_expr(column_name: str) -> pl.Expr:
    return (
        pl.col(column_name)
        .cast(pl.Utf8, strict=False)
        .str.strip_chars()
        .fill_null("UNKNOWN")
        .alias(column_name)
    )


def _empty_engine_frame() -> pl.DataFrame:
    return pl.DataFrame({column: [] for column in ENGINE_COLUMNS})


def _filter_conversion_objective(df: pl.DataFrame) -> tuple[pl.DataFrame, Dict[str, Any]]:
    objective_column = next((col for col in OBJECTIVE_CANDIDATES if col in df.columns), None)
    meta: Dict[str, Any] = {
        "objective_filter_applied": objective_column is not None,
        "objective_column": objective_column,
        "objective_value": TARGET_OBJECTIVE_VALUE,
    }
    if objective_column is None:
        return df, meta

    before_rows = int(df.height)
    filtered = df.filter(
        pl.col(objective_column).cast(pl.Utf8, strict=False).str.strip_chars().str.to_uppercase()
        == pl.lit(TARGET_OBJECTIVE_VALUE)
    )
    meta["objective_rows_before"] = before_rows
    meta["objective_rows_after"] = int(filtered.height)
    return filtered, meta


def _filter_target_divisions(df: pl.DataFrame) -> tuple[pl.DataFrame, Dict[str, Any]]:
    has_division = "DIVISION" in df.columns
    meta: Dict[str, Any] = {
        "division_filter_applied": has_division,
        "division_filter_values": list(TARGET_DIVISIONS),
    }
    if not has_division:
        return df, meta

    before_rows = int(df.height)
    filtered = df.filter(
        pl.col("DIVISION")
        .cast(pl.Utf8, strict=False)
        .str.strip_chars()
        .str.to_uppercase()
        .is_in(list(TARGET_DIVISIONS))
    )
    meta["division_rows_before"] = before_rows
    meta["division_rows_after"] = int(filtered.height)
    return filtered, meta


def _normalize_wide_engine_frame(df: pl.DataFrame) -> pl.DataFrame:
    if df.is_empty():
        return _empty_engine_frame()

    missing = sorted(set(ENGINE_COLUMNS).difference(df.columns))
    if missing:
        raise ValueError(f"Missing required columns: {missing}")

    selected = df.select(ENGINE_COLUMNS)
    _validate_metric_parse_errors(
        selected,
        metric_columns=ENGINE_METRIC_COLUMNS,
        context="wide_engine_frame",
    )

    normalized = (
        selected
        .with_columns(
            [_dimension_expr(dim) for dim in DIMENSIONS]
            + [pl.col(metric).cast(pl.Float64, strict=False).fill_null(0.0).alias(metric) for metric in ENGINE_METRIC_COLUMNS]
        )
        .select(ENGINE_COLUMNS)
    )
    return normalized.filter(pl.col("Spend_curr") > 0)


def _normalize_long_frame(df: pl.DataFrame) -> pl.DataFrame:
    year_column = "Year" if "Year" in df.columns else "YEAR" if "YEAR" in df.columns else None
    month_column = "Month" if "Month" in df.columns else "MONTH" if "MONTH" in df.columns else None
    day_column = "Day" if "Day" in df.columns else "DAY" if "DAY" in df.columns else None
    required_raw = [*DIMENSIONS, *RAW_METRIC_MAP.keys()]
    if year_column is None:
        raise ValueError("Missing required columns: ['Year']")

    missing = sorted(set(required_raw).difference(df.columns))
    if missing:
        raise ValueError(f"Missing required columns: {missing}")

    renamed = df.rename(RAW_METRIC_MAP)
    selected_columns = [*DIMENSIONS, year_column, *METRICS]
    if month_column is not None:
        selected_columns.append(month_column)
    if day_column is not None:
        selected_columns.append(day_column)

    selected = renamed.select(selected_columns)
    _validate_metric_parse_errors(
        selected,
        metric_columns=METRICS,
        context="long_engine_frame",
    )

    exprs = (
        [_dimension_expr(dim) for dim in DIMENSIONS]
        + [_metric_expr(metric) for metric in METRICS]
        + [_year_expr(year_column).alias("Year")]
    )

    output_columns = [*DIMENSIONS, "Year"]
    if month_column is not None:
        exprs.append(_int_expr(month_column).alias("Month"))
        output_columns.append("Month")
    if day_column is not None:
        exprs.append(_int_expr(day_column).alias("Day"))
        output_columns.append("Day")

    normalized = selected.with_columns(exprs)
    output_columns.extend(METRICS)
    return normalized.select(output_columns)


def _filter_target_years(df: pl.DataFrame, curr_year: int, prev_year: int) -> pl.DataFrame:
    return df.filter(pl.col("Year").is_in([curr_year, prev_year]))


def _apply_mtd_alignment(
    df: pl.DataFrame,
    curr_year: int,
    prev_year: int,
) -> tuple[pl.DataFrame, Dict[str, Any]]:
    meta: Dict[str, Any] = {
        "curr_year": curr_year,
        "prev_year": prev_year,
        "mtd_applied": False,
    }
    scoped = _filter_target_years(df, curr_year=curr_year, prev_year=prev_year)
    if scoped.is_empty():
        return scoped, meta
    if "Month" not in scoped.columns:
        return scoped, meta

    curr_scope = scoped.filter(pl.col("Year") == curr_year)
    if curr_scope.is_empty():
        return scoped, meta

    max_month = curr_scope.select(pl.col("Month").max()).to_series(0)[0]
    min_month = curr_scope.select(pl.col("Month").min()).to_series(0)[0]
    if max_month is None:
        return scoped, meta

    if min_month is not None:
        meta["mtd_month_start"] = int(min_month)
    meta["mtd_month_cutoff"] = int(max_month)

    has_day = "Day" in scoped.columns
    max_day = None
    min_day = None
    if has_day:
        if min_month is not None:
            min_day = (
                curr_scope.filter(pl.col("Month") == pl.lit(min_month))
                .select(pl.col("Day").min())
                .to_series(0)[0]
            )
        max_day = (
            curr_scope.filter(pl.col("Month") == pl.lit(max_month))
            .select(pl.col("Day").max())
            .to_series(0)[0]
        )
        if min_day is not None:
            meta["mtd_day_start"] = int(min_day)
        if max_day is not None:
            meta["mtd_day_cutoff"] = int(max_day)

    if has_day and max_day is not None:
        in_window = (pl.col("Month") < pl.lit(max_month)) | (
            (pl.col("Month") == pl.lit(max_month)) & (pl.col("Day") <= pl.lit(max_day))
        )
    else:
        in_window = pl.col("Month") <= pl.lit(max_month)

    meta["mtd_applied"] = True
    return scoped.filter(in_window), meta


def _pivot_long_to_engine(df: pl.DataFrame, curr_year: int, prev_year: int) -> pl.DataFrame:
    if df.is_empty():
        return _empty_engine_frame()

    agg_exprs: list[pl.Expr] = []
    for metric in METRICS:
        agg_exprs.append(pl.col(metric).filter(pl.col("Year") == curr_year).sum().alias(f"{metric}_curr"))
        agg_exprs.append(pl.col(metric).filter(pl.col("Year") == prev_year).sum().alias(f"{metric}_prev"))

    pivoted = df.group_by(DIMENSIONS).agg(agg_exprs).with_columns(
        [pl.col(metric).cast(pl.Float64, strict=False).fill_null(0.0).alias(metric) for metric in ENGINE_METRIC_COLUMNS]
    )

    return pivoted.select(ENGINE_COLUMNS).filter(pl.col("Spend_curr") > 0)


def _to_engine_frame(
    df: pl.DataFrame,
    curr_year: int,
    prev_year: int,
    mtd_only: bool,
) -> tuple[pl.DataFrame, Dict[str, Any]]:
    filtered_input, objective_meta = _filter_conversion_objective(df)
    scoped_input, division_meta = _filter_target_divisions(filtered_input)
    if set(ENGINE_COLUMNS).issubset(scoped_input.columns):
        meta = {
            "curr_year": curr_year,
            "prev_year": prev_year,
            "mtd_applied": False,
            "source_format": "wide",
        }
        meta.update(division_meta)
        meta.update(objective_meta)
        return _normalize_wide_engine_frame(scoped_input), meta

    long_df = _normalize_long_frame(scoped_input)
    if mtd_only:
        filtered_df, meta = _apply_mtd_alignment(long_df, curr_year=curr_year, prev_year=prev_year)
    else:
        filtered_df = _filter_target_years(long_df, curr_year=curr_year, prev_year=prev_year)
        meta = {"curr_year": curr_year, "prev_year": prev_year, "mtd_applied": False}
    meta["source_format"] = "long"
    meta.update(division_meta)
    meta.update(objective_meta)
    return _pivot_long_to_engine(filtered_df, curr_year=curr_year, prev_year=prev_year), meta


def read_input_excel(
    path: str | Path,
    preferred_sheet: str = "raw",
    curr_year: int = DEFAULT_CURR_YEAR,
    prev_year: int = DEFAULT_PREV_YEAR,
    mtd_only: bool = True,
    return_meta: bool = False,
) -> pl.DataFrame | tuple[pl.DataFrame, Dict[str, Any]]:
    """Read input Excel and return engine-ready wide schema with curr/prev metrics."""
    excel_path = Path(path)
    if not excel_path.exists():
        raise FileNotFoundError(f"Input Excel file not found: {excel_path}")

    try:
        raw_df = _read_with_polars(excel_path, preferred_sheet, preferred_sheet)
    except Exception:
        target_sheet = _select_sheet_name(excel_path, preferred_sheet)
        raw_df = _read_with_openpyxl(excel_path, target_sheet, preferred_sheet)

    engine_df, meta = _to_engine_frame(
        raw_df,
        curr_year=curr_year,
        prev_year=prev_year,
        mtd_only=mtd_only,
    )
    if return_meta:
        return engine_df, meta
    return engine_df


def _excel_cell_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, (int, bool, str)):
        return value
    if isinstance(value, float):
        if value != value or value in (float("inf"), float("-inf")):
            return None
        return value
    if isinstance(value, (list, dict)):
        return json.dumps(value, ensure_ascii=False)
    return value


def _write_with_polars(path: Path, sheets: Dict[str, pl.DataFrame]) -> bool:
    if not sheets:
        return False

    try:
        if hasattr(pl, "write_excel"):
            pl.write_excel(  # type: ignore[attr-defined]
                workbook=path,
                worksheets=sheets,
            )
            return True
    except Exception:
        pass

    first_df = next(iter(sheets.values()))
    if not hasattr(first_df, "write_excel"):
        return False

    try:
        first = True
        for sheet_name, frame in sheets.items():
            if first:
                frame.write_excel(path, worksheet=sheet_name)  # type: ignore[arg-type]
                first = False
            else:
                frame.write_excel(path, worksheet=sheet_name, mode="a")  # type: ignore[arg-type]
        return True
    except Exception:
        return False


def _write_with_openpyxl(path: Path, sheets: Dict[str, pl.DataFrame]) -> None:
    Workbook, _ = _import_openpyxl()
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    for sheet_name, frame in sheets.items():
        worksheet = workbook.create_sheet(title=str(sheet_name)[:31])
        worksheet.append(frame.columns)
        for row in frame.iter_rows(named=False):
            worksheet.append([_excel_cell_value(value) for value in row])

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def write_output_excel(path: str | Path, sheets: Dict[str, pl.DataFrame]) -> None:
    """Write output Excel with Polars-first and openpyxl fallback."""
    excel_path = Path(path)
    excel_path.parent.mkdir(parents=True, exist_ok=True)

    if _write_with_polars(excel_path, sheets):
        return
    _write_with_openpyxl(excel_path, sheets)
